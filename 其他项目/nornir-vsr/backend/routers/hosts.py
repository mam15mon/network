"""Host CRUD 接口。"""
from io import BytesIO
from typing import List, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.dependencies import get_db
from app.security import get_current_active_user
from core.db.database import Database
from core.db.models import Host
from schemas.host import (
    AddressPoolSyncResult,
    HostBatchCreate,
    HostBatchDelete,
    HostBatchEdit,
    HostCreate,
    HostOut,
    HostUpdate,
)

router = APIRouter(
    prefix="/hosts",
    tags=["hosts"],
    dependencies=[Depends(get_current_active_user)],
)


db_manager = Database()


@router.get("", response_model=List[HostOut])
def list_hosts(
    search: str | None = Query(default=None, description="搜索..."),
    site: str | None = Query(default=None, description="按站点过滤"),
    db: Session = Depends(get_db),
) -> List[HostOut]:
    query = db.query(Host)

    if search:
        pattern = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Host.name.ilike(pattern),
                Host.hostname.ilike(pattern),
                Host.username.ilike(pattern),
                Host.address_pool.ilike(pattern),
                Host.ppp_auth_mode.ilike(pattern),
                Host.site.ilike(pattern),
            )
        )

    if site:
        query = query.filter(Host.site == site.strip())

    hosts = query.order_by(Host.name).all()
    return [HostOut.model_validate(host) for host in hosts]


@router.post("", response_model=HostOut, status_code=status.HTTP_201_CREATED)
def create_host(payload: HostCreate, db: Session = Depends(get_db)) -> HostOut:
    payload_data = payload.model_dump()
    if not payload_data.get("platform"):
        payload_data["platform"] = "hp_comware"
    host = Host(**payload_data)
    try:
        db.add(host)
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Host name already exists") from exc
    db.refresh(host)
    return HostOut.model_validate(host)


@router.post("/batch")
def batch_upsert(payload: HostBatchCreate) -> dict:
    inserted, updated = db_manager.batch_add_or_update_hosts([host.model_dump() for host in payload.hosts])
    return {"inserted": inserted, "updated": updated}


@router.delete("/batch")
def batch_delete(payload: HostBatchDelete) -> dict:
    deleted = db_manager.batch_delete_hosts(payload.names)
    return {"deleted": deleted}


@router.put("/batch")
def batch_edit(payload: HostBatchEdit) -> dict:
    edited = db_manager.batch_edit_devices(payload.names, payload.data.model_dump(exclude_unset=True))
    return {"updated": edited}


@router.post("/sync-address-pools", response_model=AddressPoolSyncResult)
def sync_address_pools() -> AddressPoolSyncResult:
    """Fill host address pools by parsing latest configuration snapshots."""

    result = db_manager.sync_hosts_address_pool()
    return AddressPoolSyncResult.model_validate(result)


def _normalize_header(value: str) -> str:
    return (value or "").strip().lower()


def _map_row(headers: List[str], row: List) -> dict:
    header_map = {
        "name": {"name", "设备名称", "设备名"},
        "hostname": {"hostname", "地址", "主机名", "ip"},
        "platform": {"platform", "平台"},
        "username": {"username", "用户名"},
        "password": {"password", "密码"},
        "port": {"port", "端口"},
        "site": {"site", "站点"},
        "device_type": {"device_type", "设备类型"},
        "device_model": {"device_model", "设备型号", "型号"},
        "address_pool": {"address_pool", "地址池"},
        "ppp_auth_mode": {"ppp_auth_mode", "ppp认证模式", "认证模式"},
        "snmp_version": {"snmp_version", "snmp版本", "snmp version"},
        "snmp_community": {"snmp_community", "snmp团体字", "snmp 团体字", "团体字", "community"},
        "snmp_port": {"snmp_port", "snmp端口", "snmp port"},
    }

    header_lookup = {}
    for idx, header in enumerate(headers):
        normalized = _normalize_header(header)
        for target, aliases in header_map.items():
            if normalized in {alias.lower() for alias in aliases}:
                header_lookup[target] = idx
                break

    host_data = {}
    for field, idx in header_lookup.items():
        value = row[idx] if idx < len(row) else None
        if value is None:
            continue
        if field in {"port", "snmp_port"}:
            try:
                host_data[field] = int(value)
            except (TypeError, ValueError):
                continue
        else:
            host_data[field] = str(value).strip()

    if "platform" not in host_data or not host_data["platform"]:
        host_data["platform"] = "hp_comware"

    return host_data


@router.post("/import")
async def import_hosts(file: UploadFile = File(...)) -> dict:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持xlsx文件")

    content = await file.read()
    try:
        workbook = load_workbook(BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"读取Excel失败: {exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel内容为空")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    payloads = []
    for row in rows[1:]:
        mapped = _map_row(headers, list(row))
        if not mapped.get("name") or not mapped.get("hostname"):
            continue
        payloads.append(mapped)

    if not payloads:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="未解析到有效设备数据")

    inserted, updated = db_manager.batch_add_or_update_hosts(payloads)
    return {
        "inserted": inserted,
        "updated": updated,
        "total": len(payloads)
    }


def _export_rows() -> Tuple[List[str], List[List[str]]]:
    headers = [
        "设备名称",
        "地址",
        "平台",
        "用户名",
        "密码",
        "端口",
        "SNMP 版本",
        "SNMP 团体字",
        "SNMP 端口",
        "站点",
        "设备类型",
        "设备型号",
        "地址池",
        "PPP认证模式"
    ]
    hosts = db_manager.get_all_hosts()
    rows = []
    for device in hosts:
        rows.append([
            device.name,
            device.hostname,
            device.platform,
            device.username or "",
            device.password or "",
            str(device.port or ""),
            device.snmp_version or "",
            device.snmp_community or "",
            str(device.snmp_port or ""),
            device.site or "",
            device.device_type or "",
            device.device_model or "",
            device.address_pool or "",
            device.ppp_auth_mode or "",
        ])
    return headers, rows


@router.get("/export")
def export_hosts() -> StreamingResponse:
    headers, rows = _export_rows()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hosts"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=hosts.xlsx"
        },
    )


@router.get("/{name}", response_model=HostOut)
def get_host(name: str, db: Session = Depends(get_db)) -> HostOut:
    host = db.query(Host).filter_by(name=name).first()
    if not host:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Host not found")
    return HostOut.model_validate(host)


@router.put("/{name}", response_model=HostOut)
def update_host(name: str, payload: HostUpdate, db: Session = Depends(get_db)) -> HostOut:
    host = db.query(Host).filter_by(name=name).first()
    if not host:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Host not found")

    update_data = payload.model_dump(exclude_unset=True)
    if "name" in update_data and update_data["name"] != name:
        clash = db.query(Host).filter_by(name=update_data["name"]).first()
        if clash:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Target host name already exists")

    for key, value in update_data.items():
        setattr(host, key, value)

    db.commit()
    db.refresh(host)
    return HostOut.model_validate(host)


@router.delete("/{name}", status_code=status.HTTP_204_NO_CONTENT)
def delete_host(name: str, db: Session = Depends(get_db)) -> None:
    host = db.query(Host).filter_by(name=name).first()
    if not host:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Host not found")
    db.delete(host)
    db.commit()
